import os
from telnetlib import EC

import pytest
import time
from selenium.webdriver import ActionChains, Keys

from openpyxl.reader.excel import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pageObjects.LoginPage import LoginPage
from krishnapageObjects.ResourcesPage import Resources
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
from pageObjects.randomGen import randomGen
from GenericLib.BaseClass import BaseClass


class Test_Resources():
    baseURL = ReadConfig.getApplicationURL()
    # username = ReadConfig.getUseremail()
    password = ReadConfig.getPassword()
    # usernames2 = ReadConfig.getuseremails2()
    # usernames1 = ReadConfig.getuseremails1()
    # usernames = ReadConfig.getuseremails()

    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value
    usernames = worksheet["B6"].value
    usernames1 = worksheet["I2"].value
    usernames2 = worksheet["I3"].value
    companyname = worksheet["C2"].value

    workbook.close()
    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A2'] = username
    ws['B6'] = usernames
    ws['I2'] = usernames1
    ws['I3'] = usernames2
    ws['C2'] = companyname

    # Save the workbook
    wb.save("TestData/LoginData.xlsx")
    relative_seven = "Files/seven.jpg"
    absolute_path7 = os.path.abspath(relative_seven)
    categorydescription = "This description is Our free tool lets you easily create unique descriptions for your product pages. Just enter the product name and let our Frase AI do its magic."
    relative_eight = "Files/eight.jpg"
    absolute_path8 = os.path.abspath(relative_eight)
    subcategorydescription = "This subcategory description is Our free tool lets you easily create unique descriptions for your product pages. Just enter the product name and let our Frase AI do its magic."
    relative_one = "Files/one.png"
    absolute_path1 = os.path.abspath(relative_one)
    relative_two = "Files/two.png"
    absolute_path2 = os.path.abspath(relative_two)
    relative_three = "Files/three.png"
    absolute_path3 = os.path.abspath(relative_three)
    relative_four = "Files/four.png"
    absolute_path4 = os.path.abspath(relative_four)
    relative_five = "Files/five.png"
    absolute_path5 = os.path.abspath(relative_five)
    relative_six = "Files/six.jpg"
    absolute_path6 = os.path.abspath(relative_six)
    contentdescription = "This content description is Our free tool lets you easily create unique descriptions for your product pages. Just enter the product name and let our Frase AI do its magic."
    contentsectiondescription = "Welcome to our free and intuitive description creation tool! Simplify the process of crafting compelling product page descriptions by leveraging the power of our Frase AI. Just input the product name, and let our innovative tool work its "
    sectionimagedescription = "image preview text description presented here"
    # companyname = "all company"
    videoinput = "https://youtu.be/3lK3EyE9k4E"

    logger = LogGen.loggen()  # Logger

    @pytest.mark.test
    @pytest.mark.krishna
    @pytest.mark.regression
    @pytest.mark.run(order=64)
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_categorycreationforcompany(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_001_categorycreation **********")
        self.lp = LoginPage(driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login successful **********")

        categorytitle = randomGen.random_categorytitle()
        subcategorytitle = randomGen.random_subcategorytitle()
        contenttitle = randomGen.random_contenttitle()
        contentsectionname = randomGen.random_contentsectionname()

        self.logger.info("************** category creation test started ************")
        rs = Resources(driver)
        rs.clickoncontentmanagement()
        rs.clickoncategorynew()
        rs.setcategoryimage(self.absolute_path7)
        rs.clickoncategoryimagesave()
        rs.setcategorytitle(categorytitle)
        rs.setcategorydescription(self.categorydescription)
        rs.clickoncategorypublic()
        rs.clickoncategoryenable()
        rs.clickoncategorysave()
        # time.sleep(3)
        self.logger.info("************** category creation test completed ************")

        xpath = "//div[contains(text(),'Category created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_category_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(2)
        self.logger.info("************** subcategory creation test started ************")
        rs.setcategorysearch(categorytitle)
        rs.clickClosetoaster()
        rs.clickoncategoryclick()
        rs.clickonsubcategorynew()
        rs.clickonsubcategorybutton()
        rs.setsubcategoryimage(self.absolute_path8)
        rs.clickoncategoryimagesave()
        rs.setsubcategorytitle(subcategorytitle)
        rs.setcategorydescription(self.subcategorydescription)
        rs.clickonsubcategoryenable()
        rs.clickonsubcategorysave()
        self.logger.info("************** subcategory creation test completed ************")

        xpath = "//div[contains(text(),'Subcategory created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_subcategory_creation.png")
            driver.close()
            driver.quit()
            assert False

        self.logger.info("********** content creation test is started *********")
        rs.clickonsubcategorynew()
        rs.clickoncontentnew()
        rs.clickonaddcategory()
        rs.clickClosetoaster()
        rs.clickoncategoryclose()
        rs.setcontentbannerimage(self.absolute_path1)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path2)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path3)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path4)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path5)
        rs.clickonbannerimagesave()
        rs.setcontenttitle(contenttitle)
        rs.setcontentdescription(self.contentdescription)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        # rs.scrollpublic()
        rs.clickoncontentcanshare()
        rs.setcontentsectionname(contentsectionname)
        rs.setcontentsectiondescription(self.contentsectiondescription)
        rs.clickonuploadvideo()
        rs.setvideoinput(self.videoinput)
        rs.clickonuploadvideofile()
        rs.clickonaddvideo()
        # rs.clickonsectionimageselect()
        # rs.setsectionimageselect(self.absolute_path1)
        # rs.setsectionimagedescription(self.sectionimagedescription)
        rs.clickonsectionsave()
        rs.clickoncontentpublish()
        xpath = "//div[contains(text(),'Content created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_content_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(3)
        rs.clickonresources()
        rs.setresourcescategorysearch(categorytitle)
        time.sleep(2)

        driver.find_element(By.XPATH, "(//span[text()='" + categorytitle + "'])[1]").click()
        time.sleep(3)
        if "Videos" in driver.page_source:
            self.logger.info("********** content verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_contentverificationinmyresources.png")
            assert False
        time.sleep(3)
        rs.clickonbackresources()
        time.sleep(1)
        rs.setresourcescategorysearch(categorytitle)
        time.sleep(2)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        time.sleep(2)
        driver.find_element(By.XPATH, "(//span[text()='" + categorytitle + "'])[2]").click()
        time.sleep(3)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_contentverificationinmyresources.png")
            assert False
        time.sleep(3)
        rs.clickonlogout()
        self.logger.info("************* Logout successful **********")
        self.lp.setUserName(self.usernames1)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* relationcompanyLogin successful **********")
        rs.clickonresources()
        rs.clickonnetworkresources()
        rs.setsearchcompanyname(self.companyname)
        rs.clickoncompanyselect()
        rs.clickoncontents()
        rs.setsearchcontents(contenttitle)
        time.sleep(2)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        time.sleep(1)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")


        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categorycreationforcompany.png")
            assert False
        time.sleep(3)
        rs.clickoncontentclose()
        rs.clickonvideosbutton()
        rs.clickonvideoselect()
        rs.clickonvideoshare()
        time.sleep(3)
        if "Share" in driver.page_source:
            self.logger.info("********** content verification test is passed *********")


        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categorycreationforcompany.png")
            assert False
        time.sleep(3)

    @pytest.mark.sanity
    @pytest.mark.babi
    @pytest.mark.run(order=65)
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_categorycreationforemployee(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_002_categorycreation **********")
        driver.implicitly_wait(20)
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login successful **********")

        categorytitle = randomGen.random_categorytitle()
        subcategorytitle = randomGen.random_subcategorytitle()
        contenttitle = randomGen.random_contenttitle()
        contentsectionname = randomGen.random_contentsectionname()

        self.logger.info("************** category creation test started ************")
        rs = Resources(driver)
        rs.clickoncontentmanagement()
        rs.clickoncategorynew()
        rs.setcategoryimage(self.absolute_path7)
        rs.clickoncategoryimagesave()
        rs.setcategorytitle(categorytitle)
        rs.setcategorydescription(self.categorydescription)
        rs.clickoncategorypublic()
        rs.clickoncategoryenable()
        rs.clickoncategorysave()
        time.sleep(3)
        self.logger.info("************** category creation test completed ************")

        xpath = "//div[contains(text(),'Category created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_category_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(2)
        self.logger.info("************** subcategory creation test started ************")
        rs.setcategorysearch(categorytitle)
        rs.clickoncategoryclick()
        rs.clickonsubcategorynew()
        rs.clickonsubcategorybutton()
        rs.setsubcategoryimage(self.absolute_path8)
        rs.clickoncategoryimagesave()
        rs.setsubcategorytitle(subcategorytitle)
        rs.setcategorydescription(self.subcategorydescription)
        rs.clickonsubcategoryenable()
        rs.clickonsubcategorysave()
        self.logger.info("************** subcategory creation test completed ************")
        xpath = "//div[contains(text(),'Subcategory created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_subcategory.png")
            driver.close()
            driver.quit()
            assert False

        self.logger.info("********** content creation test is started *********")
        rs.clickonsubcategorynew()
        rs.clickoncontentnew()
        rs.clickonaddcategory()
        rs.clickClosetoaster()
        rs.clickoncategoryclose()
        rs.setcontentbannerimage(self.absolute_path1)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path2)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path3)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path4)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path5)
        rs.clickonbannerimagesave()
        rs.setcontenttitle(contenttitle)
        rs.setcontentdescription(self.contentdescription)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        rs.clickoncontentcanshare()
        rs.setcontentsectionname(contentsectionname)
        rs.setcontentsectiondescription(self.contentsectiondescription)
        rs.clickonsectionsave()
        rs.clickoncontentpublish()
        xpath = "//div[contains(text(),'Content created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_content_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(3)
        rs.clickonresources()
        time.sleep(3)
        rs.setresourcescategorysearch(categorytitle)
        time.sleep(2)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        time.sleep(1)
        driver.find_element(By.XPATH, "//span[text()='" + categorytitle + "']").click()
        time.sleep(3)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_contentverificationinmyresources.png")
            assert False

        time.sleep(3)
        rs.clickonlogout()
        self.logger.info("************* Logout successful **********")
        time.sleep(3)
        lp.setUserName(self.usernames)
        lp.setPassword(self.password)
        lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* relationcompanyLogin successful **********")
        rs.clickonresources()
        time.sleep(3)
        rs.setresourcescategorysearch(categorytitle)
        time.sleep(2)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        time.sleep(1)
        driver.find_element(By.XPATH, "//span[text()='" + categorytitle + "']").click()
        time.sleep(3)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categorycreationforemployee.png")
            assert False
        time.sleep(3)

    @pytest.mark.test
    @pytest.mark.regression
    @pytest.mark.run(order=66)
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_categorycreationforhierarchy(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_003_categorycreation **********")
        driver.implicitly_wait(20)
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login successful **********")

        categorytitle = randomGen.random_categorytitle()
        subcategorytitle = randomGen.random_subcategorytitle()
        subcategorytitle1 = randomGen.random_subcategorytitle1()
        subcategorytitle2 = randomGen.random_subcategorytitle2()
        subcategorytitle3 = randomGen.random_subcategorytitle3()
        subcategorytitle4 = randomGen.random_subcategorytitle4()

        self.logger.info("************** category creation test started ************")
        rs = Resources(driver)
        rs.clickoncontentmanagement()
        rs.clickoncategorynew()
        rs.setcategoryimage(self.absolute_path7)
        rs.clickoncategoryimagesave()
        rs.setcategorytitle(categorytitle)
        rs.setcategorydescription(self.categorydescription)
        rs.clickoncategorypublic()
        rs.clickoncategoryenable()
        rs.clickoncategorysave()
        time.sleep(3)
        self.logger.info("************** category creation test completed ************")

        xpath = "//div[contains(text(),'Category created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_category_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(2)
        self.logger.info("************** subcategory creation test started ************")
        rs.setcategorysearch(categorytitle)
        rs.clickoncategoryclick()
        rs.clickonsubcategorynew()
        rs.clickonsubcategorybutton()
        rs.setsubcategoryimage(self.absolute_path8)
        rs.clickoncategoryimagesave()
        rs.setsubcategorytitle(subcategorytitle)
        rs.setcategorydescription(self.subcategorydescription)
        rs.clickonsubcategoryenable()
        rs.clickonsubcategorysave()
        self.logger.info("************** subcategory creation test completed ************")

        xpath = "//div[contains(text(),'Subcategory created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_subcategory_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(3)
        subcategories = [subcategorytitle, subcategorytitle1, subcategorytitle2, subcategorytitle3, subcategorytitle4]
        for i in range(len(subcategories)):
            self.logger.info(f"************** subcategory{i + 1} creation test started ************")
            rs.setcategorysearch(subcategories[i - 1] if i > 0 else categorytitle)
            time.sleep(3)
            rs.clickonsubcategoryclick()
            rs.clickonsubcategorynew()
            rs.clickonsubcategorybutton()
            rs.setsubcategoryimage(self.absolute_path6)
            rs.clickoncategoryimagesave()
            rs.setsubcategorytitle(subcategories[i])
            rs.setcategorydescription(self.subcategorydescription)
            rs.clickonsubcategoryenable()
            rs.clickonsubcategorysave()
            self.logger.info(f"************** subcategory{i + 1} creation test completed ************")

            xpath = "//div[contains(text(),'Subcategory created successfully')]"
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )

            if element:
                self.logger.info(f"Found Toast message : {element.text}")
                assert True
            else:
                self.logger.info(f"Toast message not found: {element.text}")
                driver.save_screenshot(f".\\Screenshots\\test_subcategory_creation_{i + 1}.png")
                driver.close()
                driver.quit()
                assert False

        time.sleep(3)
        self.logger.info("************** content verification test started ************")
        rs.setcategorysearch(subcategorytitle4)
        time.sleep(3)
        if "You have reached the last subcategory, cannot go inside" in driver.page_source:
            self.logger.info("********** content verification test is passed *********")
        else:
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categorycreationforhierarchy.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=67)
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_categorycreationforrelationcompany(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_004_categorycreation **********")
        self.lp = LoginPage(driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login successful **********")

        categorytitle = randomGen.random_categorytitle()
        subcategorytitle = randomGen.random_subcategorytitle()
        contenttitle = randomGen.random_contenttitle()
        contentsectionname = randomGen.random_contentsectionname()

        self.logger.info("************** category creation test started ************")
        rs = Resources(driver)
        rs.clickoncontentmanagement()
        rs.clickoncategorynew()
        rs.setcategoryimage(self.absolute_path7)
        rs.clickoncategoryimagesave()
        rs.setcategorytitle(categorytitle)
        rs.setcategorydescription(self.categorydescription)
        rs.clickoncategorypartner()
        rs.clickoncategoryenable()
        rs.clickoncategorysave()
        time.sleep(3)
        self.logger.info("************** category creation test completed ************")

        xpath = "//div[contains(text(),'Category created successfully')]"
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_category_creation.png")
            assert False

        time.sleep(2)
        self.logger.info("************** subcategory creation test started ************")
        rs.setcategorysearch(categorytitle)
        rs.clickoncategoryclick()
        rs.clickonsubcategorynew()
        rs.clickonsubcategorybutton()
        rs.setsubcategoryimage(self.absolute_path8)
        rs.clickoncategoryimagesave()
        rs.setsubcategorytitle(subcategorytitle)
        rs.setcategorydescription(self.subcategorydescription)
        rs.clickonsubcategoryenable()
        rs.clickonsubcategorysave()
        self.logger.info("************** subcategory creation test completed ************")
        xpath = "//div[contains(text(),'Subcategory created successfully')]"
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_subcategory_creation.png")
            assert False

        self.logger.info("********** content creation test is started *********")
        rs.clickonsubcategorynew()
        rs.clickoncontentnew()
        rs.clickonaddcategory()
        rs.clickClosetoaster()
        rs.clickoncategoryclose()
        rs.setcontentbannerimage(self.absolute_path1)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path2)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path3)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path4)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path5)
        rs.clickonbannerimagesave()
        rs.setcontenttitle(contenttitle)
        rs.setcontentdescription(self.contentdescription)
        actions = ActionChains(driver)
        actions.send_keys(Keys.PAGE_DOWN)
        actions.perform()
        rs.clickoncontentcanshare()
        rs.setcontentsectionname(contentsectionname)
        rs.setcontentsectiondescription(self.contentsectiondescription)
        rs.clickonsectionsave()
        rs.clickoncontentpublish()
        xpath = "//div[contains(text(),'Content created successfully')]"
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_content_creation.png")
            assert False

        time.sleep(3)
        rs.clickonresources()
        time.sleep(3)
        rs.setresourcescategorysearch(categorytitle)
        time.sleep(2)
        actions = ActionChains(driver)
        actions.send_keys(Keys.PAGE_DOWN)
        actions.perform()
        time.sleep(1)
        driver.find_element(By.XPATH, "//span[text()='" + categorytitle + "']").click()
        time.sleep(3)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")
        else:
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categoryverificationinmyresources.png")
            assert False

        time.sleep(3)
        rs.clickonlogout()
        self.logger.info("************* Logout successful **********")
        time.sleep(3)
        self.lp.setUserName(self.usernames2)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* relationcompanyLogin successful **********")
        rs.clickonresources()
        time.sleep(3)
        rs.clickonnetworkresources()
        time.sleep(3)
        rs.setsearchcompanyname(self.companyname)
        time.sleep(3)
        rs.clickoncompanyselect()
        time.sleep(3)
        rs.clickoncontents()
        time.sleep(3)
        rs.setsearchcontents(contenttitle)
        time.sleep(3)
        actions = ActionChains(driver)
        actions.send_keys(Keys.PAGE_DOWN)
        actions.perform()
        time.sleep(1)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")
        else:
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categorycreationforrelationcompany.png")
            assert False
        time.sleep(3)

    @pytest.mark.sohl
    @pytest.mark.regression
    @pytest.mark.run(order=68)
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_categoryshare(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_005_categorycreation **********")
        driver.implicitly_wait(20)
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login succesful **********")

        categorytitle = randomGen.random_categorytitle()
        subcategorytitle = randomGen.random_subcategorytitle()
        contenttitle = randomGen.random_contenttitle()
        contentsectionname = randomGen.random_contentsectionname()

        self.logger.info("************** category creation test started ************")
        rs = Resources(driver)
        rs.clickoncontentmanagement()
        rs.clickoncategorynew()
        rs.setcategoryimage(self.absolute_path7)
        rs.clickoncategoryimagesave()
        rs.setcategorytitle(categorytitle)
        rs.setcategorydescription(self.categorydescription)
        rs.clickoncategorypublic()
        rs.clickoncategoryenable()
        rs.clickoncategorysave()
        time.sleep(3)
        self.logger.info("************** category creation test completed ************")

        xpath = "//div[contains(text(),'Category created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_category_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(2)
        self.logger.info("************** subcategory creation test started ************")
        rs.setcategorysearch(categorytitle)
        rs.clickoncategoryclick()
        rs.clickonsubcategorynew()
        rs.clickonsubcategorybutton()
        rs.setsubcategoryimage(self.absolute_path8)
        rs.clickoncategoryimagesave()
        rs.setsubcategorytitle(subcategorytitle)
        rs.setcategorydescription(self.subcategorydescription)
        rs.clickonsubcategoryenable()
        rs.clickonsubcategorysave()
        self.logger.info("************** subcategory creation test completed ************")

        xpath = "//div[contains(text(),'Subcategory created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_subcategory_creation.png")
            driver.close()
            driver.quit()
            assert False

        self.logger.info("********** content creation test is started *********")
        rs.clickonsubcategorynew()
        rs.clickoncontentnew()
        rs.clickonaddcategory()
        rs.clickClosetoaster()
        rs.clickoncategoryclose()
        rs.setcontentbannerimage(self.absolute_path1)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path2)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path3)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path4)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path5)
        rs.clickonbannerimagesave()
        rs.setcontenttitle(contenttitle)
        rs.setcontentdescription(self.contentdescription)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        rs.clickoncontentcanshare()
        rs.setcontentsectionname(contentsectionname)
        rs.setcontentsectiondescription(self.contentsectiondescription)
        rs.clickonsectionsave()
        rs.clickoncontentpublish()
        xpath = "//div[contains(text(),'Content created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_content_creation.png")
            driver.close()
            driver.quit()
            assert False

        rs.clickonresources()
        time.sleep(3)
        rs.setresourcescategorysearch(categorytitle)
        time.sleep(2)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        time.sleep(1)
        driver.find_element(By.XPATH, "//span[text()='" + categorytitle + "']").click()
        time.sleep(3)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categoryverificationinmyresources.png")
            assert False
        time.sleep(3)
        rs.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        time.sleep(3)
        lp.setUserName(self.usernames1)
        lp.setPassword(self.password)
        lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* relationcompanyLogin succesful **********")
        rs.clickonresources()
        time.sleep(3)
        rs.clickonnetworkresources()
        time.sleep(3)
        rs.setsearchcompanyname(self.companyname)
        time.sleep(3)
        rs.clickoncompanyselect()
        time.sleep(3)
        rs.clickoncontents()
        time.sleep(3)
        rs.setsearchcontents(contenttitle)
        time.sleep(3)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        time.sleep(1)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categoryforrelationcompany.png")
            assert False
        time.sleep(3)
        rs.clickonlogout()
        self.logger.info("************* Logout succesful **********")
        time.sleep(3)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login succesful **********")
        rs.clickoncontentmanagement()
        time.sleep(3)
        rs.setcategorysearch(categorytitle)
        time.sleep(3)
        rs.clickoncategoryshare()
        rs.clickoncategorysharepublic()
        rs.clickoncategoryshareaccessyes()
        rs.clickoncategorypartner()
        rs.clickoncategoryshareaccessyes()
        time.sleep(3)
        rs.clickoncategoryclose()
        rs.clickonlogout()
        time.sleep(3)
        lp.setUserName(self.usernames2)
        lp.setPassword(self.password)
        lp.clickLogin()
        time.sleep(3)
        self.logger.info("************* relationcompanyLogin succesful **********")
        rs.clickonresources()
        time.sleep(3)
        rs.clickonnetworkresources()
        time.sleep(3)
        rs.setsearchcompanyname(self.companyname)
        time.sleep(3)
        rs.clickoncompanyselect()
        time.sleep(3)
        rs.clickoncontents()
        time.sleep(3)
        rs.setsearchcontents(contenttitle)
        time.sleep(3)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        time.sleep(1)
        driver.find_element(By.XPATH, "//h4[normalize-space()='" + contenttitle + "']").click()
        time.sleep(3)
        if "This content description is Our free tool lets you easily create unique descriptions for your product pages." in driver.page_source:
            self.logger.info("********** content verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** content verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_categoryforrelationcompany.png")
            assert False
        time.sleep(3)
        rs.clickonlogout()
        self.logger.info("************* Logout succesful **********")

    @pytest.mark.test
    @pytest.mark.regression
    @pytest.mark.run(order=69)
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_categoryedit(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_006_categorycreation **********")
        driver.implicitly_wait(20)
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login successful **********")

        categorytitle = randomGen.random_categorytitle()
        subcategorytitle = randomGen.random_subcategorytitle()
        contenttitle = randomGen.random_contenttitle()
        contentsectionname = randomGen.random_contentsectionname()

        self.logger.info("************** category creation test started ************")
        rs = Resources(driver)
        rs.clickoncontentmanagement()
        rs.clickoncategorynew()
        rs.setcategoryimage(self.absolute_path7)
        rs.clickoncategoryimagesave()
        rs.setcategorytitle(categorytitle)
        rs.setcategorydescription(self.categorydescription)
        rs.clickoncategorypublic()
        rs.clickoncategoryenable()
        rs.clickoncategorysave()
        time.sleep(3)
        self.logger.info("************** category creation test completed ************")

        xpath = "//div[contains(text(),'Category created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_category_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(2)
        self.logger.info("************** subcategory creation test started ************")
        rs.setcategorysearch(categorytitle)
        rs.clickoncategoryclick()
        rs.clickonsubcategorynew()
        rs.clickonsubcategorybutton()
        rs.setsubcategoryimage(self.absolute_path8)
        rs.clickoncategoryimagesave()
        rs.setsubcategorytitle(subcategorytitle)
        rs.setcategorydescription(self.subcategorydescription)
        rs.clickonsubcategoryenable()
        rs.clickonsubcategorysave()
        self.logger.info("************** subcategory creation test completed ************")
        xpath = "//div[contains(text(),'Subcategory created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_subcategory_creation.png")
            driver.close()
            driver.quit()
            assert False

        self.logger.info("********** content creation test is started *********")
        rs.clickonsubcategorynew()
        rs.clickoncontentnew()
        rs.clickonaddcategory()
        rs.clickClosetoaster()
        rs.clickoncategoryclose()
        rs.setcontentbannerimage(self.absolute_path1)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path2)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path3)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path4)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path5)
        rs.clickonbannerimagesave()
        rs.setcontenttitle(contenttitle)
        rs.setcontentdescription(self.contentdescription)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        rs.clickoncontentcanshare()
        rs.setcontentsectionname(contentsectionname)
        rs.setcontentsectiondescription(self.contentsectiondescription)
        rs.clickonsectionsave()
        rs.clickoncontentpublish()
        xpath = "//div[contains(text(),'Content created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_content_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(3)
        rs.clickoncontentmanagementbreadcrumb()
        time.sleep(3)
        rs.setcategorysearch(categorytitle)
        rs.clickoncategoryclick()
        rs.clickoncategoryedit()
        rs.clickoncategorydisable()
        rs.clickoncategoryupdate()
        xpath = "//div[contains(text(),'Category updated successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_Category_updated.png")
            driver.close()
            driver.quit()
            assert False

        rs.scrollcontent()
        rs.clickonsubcategoryedit()
        rs.clickonsubcategorydisable()
        rs.clickonsubcategoryupdate()
        xpath = "//div[contains(text(),'Subcategory updated successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_Subcategory_updated.png")
            driver.close()
            driver.quit()
            assert False

        rs.clickoncontentviewmore()
        rs.clickonsectionedit()
        rs.setcontentsectionname(contentsectionname)
        rs.clickonsectionupdate()
        xpath = "//div[contains(text(),'Section updated successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_Section_updated.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(3)

    # @pytest.mark.test
    @pytest.mark.regression
    @pytest.mark.run(order=70)
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_categorydelete(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_007_categorycreation **********")
        self.lp = LoginPage(driver)
        self.lp.setUserName(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.logger.info("************* Login successful **********")

        categorytitle = randomGen.random_categorytitle()
        subcategorytitle = randomGen.random_subcategorytitle()
        contenttitle = randomGen.random_contenttitle()
        contentsectionname = randomGen.random_contentsectionname()

        self.logger.info("************** category creation test started ************")
        rs = Resources(driver)
        rs.clickoncontentmanagement()
        rs.clickoncategorynew()
        rs.setcategoryimage(self.absolute_path7)
        rs.clickoncategoryimagesave()
        rs.setcategorytitle(categorytitle)
        rs.setcategorydescription(self.categorydescription)
        rs.clickoncategorypublic()
        rs.clickoncategoryenable()
        rs.clickoncategorysave()
        time.sleep(3)
        self.logger.info("************** category creation test completed ************")

        xpath = "//div[contains(text(),'Category created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_category_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(2)
        self.logger.info("************** subcategory creation test started ************")
        rs.setcategorysearch(categorytitle)
        rs.clickoncategoryclick()
        rs.clickonsubcategorynew()
        rs.clickonsubcategorybutton()
        rs.setsubcategoryimage(self.absolute_path8)
        rs.clickoncategoryimagesave()
        rs.setsubcategorytitle(subcategorytitle)
        rs.setcategorydescription(self.subcategorydescription)
        rs.clickonsubcategoryenable()
        rs.clickonsubcategorysave()
        self.logger.info("************** subcategory creation test completed ************")
        xpath = "//div[contains(text(),'Subcategory created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_subcategory_creation.png")
            driver.close()
            driver.quit()
            assert False

        self.logger.info("********** content creation test is started *********")
        rs.clickonsubcategorynew()
        rs.clickoncontentnew()
        rs.clickonaddcategory()
        rs.clickClosetoaster()
        rs.clickoncategoryclose()
        rs.setcontentbannerimage(self.absolute_path1)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path2)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path3)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path4)
        rs.clickonbannerimagesave()
        rs.setcontentbannerimage(self.absolute_path5)
        rs.clickonbannerimagesave()
        rs.setcontenttitle(contenttitle)
        rs.setcontentdescription(self.contentdescription)
        actions = ActionChains(driver)

        # Press the PAGE_DOWN key to scroll down
        actions.send_keys(Keys.PAGE_DOWN)

        # Perform the scrolling action
        actions.perform()
        rs.clickoncontentcanshare()
        rs.setcontentsectionname(contentsectionname)
        rs.setcontentsectiondescription(self.contentsectiondescription)
        rs.clickonsectionsave()
        rs.clickoncontentpublish()
        xpath = "//div[contains(text(),'Content created successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_content_creation.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(3)
        rs.clickoncontentmanagementbreadcrumb()
        rs.setcategorysearch(categorytitle)
        rs.clickoncategoryclick()
        rs.scrollcontent()
        rs.clickonsubcategorydelete()
        rs.clickoncontentconfirmdelete()
        xpath = "//div[contains(text(),'Subcategory deleted successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_Subcategory_deleted.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(3)
        rs.scrollcontent()
        rs.clickoncontentviewmore()
        rs.clickoncontentdelete()
        rs.clickoncontentconfirmdelete()
        xpath = "//div[contains(text(),'Content and its sections deleted successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_content_deleted.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(3)
        rs.clickonsubcategorydelete()
        rs.clickoncontentconfirmdelete()
        xpath = "//div[contains(text(),'Category deleted successfully')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Toast message : {element.text}")
            assert True
        else:
            self.logger.info(f"Toast message not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_Category_deleted.png")
            driver.close()
            driver.quit()
            assert False

        time.sleep(2)

    if __name__ == '__main__':
        unittest.main(verbosity=2)
