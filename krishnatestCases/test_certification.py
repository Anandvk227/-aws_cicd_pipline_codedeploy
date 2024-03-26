import os

import pytest
import time

from openpyxl.reader.excel import load_workbook
from selenium.common import ElementNotInteractableException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from pageObjects.LoginPage import LoginPage
from krishnapageObjects.CertificationPage import Certification
from pageObjects.randomGen import randomGen
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
from GenericLib.BaseClass import BaseClass

class Test_Certification():
    baseURL = ReadConfig.getApplicationURL()
    # username = ReadConfig.getUseremail()
    # usernames4 = ReadConfig.getuseremails4()
    # usernames5 = ReadConfig.getuseremails5()
    password = ReadConfig.getPassword()
    enterquestion = "who  is the cm of Telangana"
    firstanswer = "KCR"
    secondanswer = "Revanth Reddy"
    templatename = "General Knowledge"
    relative_five = "Files/five.png"
    absolute_path5 = os.path.abspath(relative_five)
    programname = "Gk General Knowledge"
    categoryname = "category"
    relative_three = "Files/three.png"
    absolute_path3 = os.path.abspath(relative_three)
    signname = "krishna"
    signdesignation = "QA"
    certificationdescription = "Your certification is typically displayed as a document stating that as a professional, you've been trained, educated and are prepared to meet a specific set of criteria for your role."
    questionnumber = "1"
    marks = "80"
    reapply = "1"
    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value
    usernames4 = worksheet["I2"].value
    usernames5 = worksheet["B6"].value

    workbook.close()
    # Load the existing workbook
    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A2'] = username
    ws['I2'] = usernames4
    ws['B6'] = usernames5

    # Save the workbook
    wb.save("TestData/LoginData.xlsx")

    logger = LogGen.loggen()  # Logger

    @pytest.mark.tests
    @pytest.mark.regression
    @pytest.mark.run(order=71)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_certificationcreatingeditinganddeleting(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_001_Certification **********")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login successful **********")
        acronym1 = randomGen.random_acronym1()
        acronym2 = randomGen.random_acronym2()
        certificationname = randomGen.random_certificationname()
        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A21'] = certificationname

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.logger.info("******* Starting acronym Test **********")
        cp = Certification(driver)
        cp.clickoncertificationprogramme()
        cp.clickonmarkingsystem()
        cp.clickonmarkingsystemnew()
        cp.setacronym1(acronym1)
        cp.clickonaddanotherfield()
        cp.setacronym2(acronym2)
        cp.clickonsave()
        self.logger.info(
            "*********** TC_01 Create the marking system both published and unpublished list ***************")

        xpath = "//div[contains(text(),'Acronym created and unpublished successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_acronymcreation.png")
            assert False
        time.sleep(3)
        cp.setsearch(acronym1)
        cp.clickonacronymedit()
        self.logger.info("*********** TC_02 move the unpublished to publish. the marking System  **********")
        cp.clickonacronympublish()
        self.logger.info("************ TC_03 edit and delete the marking system ***************")

        xpath = "//div[contains(text(),'Acronym updated and published successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_acronymupdate.png")
            assert False
        time.sleep(3)
        cp.clickonquestionbank()
        cp.clickonmarkingsystemnew()
        cp.setenterquestion(self.enterquestion)
        cp.setfirstanswer(self.firstanswer)
        cp.clickonadd()
        cp.setsecondanswer(self.secondanswer)
        cp.clickonselectanswer()
        cp.clickonselectmarkingsystem()
        cp.clickonselectmarkingsystemoption()
        cp.clickonselectacronym()
        cp.clickonquestionsave()
        time.sleep(3)
        if "Question created and unpublished successfully" in driver.page_source:
            self.logger.info("********** Question creation test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Question creation test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_questioncreation.png")
            assert False

        xpath = "//div[contains(text(),'Question created and unpublished successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_questioncreation.png")
            assert False

        cp = ConfigurationPage(driver)
        cp.clickonpublishedtab()
        cp.clickonunpublishedtab()
        time.sleep(2)
        cp.setsearch(self.enterquestion)
        cp.clickonquestionedit()
        cp.clickonbuttonedit()
        cp.clickonquestionpublish()

        xpath = "//div[contains(text(),'Question update and published successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_questionupdate.png")
            assert False

        cp.clickontemplates()
        cp.clickonmarkingsystemnew()
        time.sleep(1)
        hover_element = driver.find_element(By.XPATH, "//div[@class='templatemediaOverlay certHover']")
        action = ActionChains(driver)
        action.move_to_element(hover_element).perform()

        edit_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH,
                                        "/html[1]/body[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/button[1]"))
        )
        edit_button.click()
        cp.settemplatename(self.templatename)
        cp.setlogoimg(self.absolute_path5)
        cp.clickonimgsave()
        cp.setprogramname(self.programname)
        cp.setcategoryname(self.categoryname)
        cp.setsign(self.absolute_path3)
        cp.clickonimgsave()
        cp.setsignname(self.signname)
        cp.setsigndesignation(self.signdesignation)
        cp.clickontemplatesave()

        xpath = "//div[contains(text(),'Created and unpublished successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_templatecreation.png")
            assert False

        cp.clickonedittemplate()
        time.sleep(1)
        cp.scrollnote()
        cp.clickontemplatepublish()

        xpath = "//div[contains(text(),'Updated and published successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_templateupdate.png")
            assert False

        cp.clickoncertification()
        cp.clickonmarkingsystemnew()
        cp.setcertificationname(certificationname)
        cp.setcertificationdescription(self.certificationdescription)
        cp.scrollupto()
        cp.clickonpublic()
        cp.clickonselectmarkingsystemfield()
        cp.clickonacronymselect()
        cp.clickonquestionnumber()
        cp.setquestionnumber(self.questionnumber)
        cp.setMarks(self.marks)
        cp.setreapply(self.reapply)
        cp.clickonminutes()
        cp.clickonminutestime()
        cp.clickonoptions()
        cp.clickonselectoption()
        cp.clickonselecttemplate()
        cp.clickontemplateclick()
        cp.clickoncertificatesave()
        cp.clickoncertificateconfirmsave()

        xpath = "//div[contains(text(),'Created and unpublished successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificationcreation.png")
            assert False

        cp.setsearch(certificationname)
        time.sleep(2)
        driver.find_element(By.XPATH, "//span[text()='" + certificationname + "']").click()
        cp.clickoncertificateedit()
        cp.clickoncertificatepublish()
        cp.clickoncertificateconfirmpublish()

        xpath = "//div[contains(text(),'Updated and published successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificateupdate.png")
            assert False

        cp.setsearch(certificationname)
        time.sleep(2)
        wait = WebDriverWait(driver, 10)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, "//span[text()='" + certificationname + "']")))
        element.click()
        cp.clickoncertificatedelete()
        cp.clickoncertificateconfirmdelete()
        xpath = "//div[contains(text(),'Deleted successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificatedelete.png")
            assert False

        cp.clickontemplates()
        cp.clickontemplatedelete()
        cp.clickoncertificateconfirmdelete()

        xpath = "//div[contains(text(),'Deleted successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificatedelete.png")
            assert False

        cp.clickonquestionbank()
        cp.clickonquestionedit()
        cp.clickonquestionbankdelete()
        cp.clickoncertificateconfirmdelete()

        xpath = "//div[contains(text(),'Deleted successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificatedelete.png")
            assert False

        cp.clickonmarkingsystem()
        cp.clickonacronymdelete()
        cp.clickoncertificateconfirmdelete()
        xpath = "//div[contains(text(),'Acronym deleted successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_acronymdelete.png")
            assert False

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=72)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_certificationcreating(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_001_Certification **********")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login succesful **********")
        acronym1 = randomGen.random_acronym1()
        acronym2 = randomGen.random_acronym2()
        certificationname = randomGen.random_certificationname()
        self.logger.info("******** Generating and storing data into excel sheet ***********")
        # Load the existing workbook
        wb = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        ws = wb.active

        # Update the existing cells with new data
        ws['A21'] = certificationname

        # Save the workbook
        wb.save("TestData/LoginData.xlsx")

        self.logger.info("******* Starting acronym Test **********")
        cp = Certification(driver)
        cp.clickoncertificationprogramme()
        cp.clickonmarkingsystem()
        cp.clickonmarkingsystemnew()
        cp.setacronym1(acronym1)
        cp.clickonaddanotherfield()
        cp.setacronym2(acronym2)

        cp.clickonsave()
        self.logger.info(
            "*********** TC_01 Create the marking sysyem both published and unpublished list ***************")

        xpath = "//div[contains(text(),'Acronym created and unpublished successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_acronymcreation.png")
            assert False
        time.sleep(3)
        cp.setsearch(acronym1)
        cp.clickonacronymedit()
        self.logger.info("*********** TC_02 move the unpublish to publish. the marking System  **********")
        cp.clickonacronympublish()
        self.logger.info("************ TC_03 edit and delete the marking system ***************")

        xpath = "//div[contains(text(),'Acronym updated and published successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_acronymupdate.png")
            assert False
        time.sleep(3)
        cp.clickonquestionbank()
        cp.clickonmarkingsystemnew()
        cp.setenterquestion(self.enterquestion)
        cp.setfirstanswer(self.firstanswer)
        cp.clickonadd()
        cp.setsecondanswer(self.secondanswer)
        cp.clickonselectanswer()
        cp.clickonselectmarkingsystem()
        cp.clickonselectmarkingsystemoption()
        cp.clickonselectacronym()
        cp.clickonquestionsave()
        time.sleep(3)
        if "Question created and unpublished successfully" in driver.page_source:
            self.logger.info("********** Question creation test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Question creation test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_questioncreation.png")
            assert False
        xpath = "//div[contains(text(),'Question created and unpublished successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_questioncreation.png")
            assert False
        # time.sleep(3)
        cp.clickonpublishedtab()
        cp.clickonunpublishedtab()
        time.sleep(2)
        cp.setsearch(self.enterquestion)
        cp.clickonquestionedit()
        cp.clickonbuttonedit()
        cp.clickonquestionpublish()

        xpath = "//div[contains(text(),'Question update and published successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_questionupdate.png")
            assert False
        time.sleep(3)
        cp.clickontemplates()
        cp.clickonmarkingsystemnew()
        time.sleep(1)
        # Locate the element to perform mouse hover
        hover_element = driver.find_element(By.XPATH, "//div[@class='templatemediaOverlay certHover']")

        # Perform mouse hover action
        action = ActionChains(driver)
        action.move_to_element(hover_element).perform()

        # Locate and click on the edit button
        edit_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH,
                                        "/html[1]/body[1]/div[1]/div[1]/div[1]/div[2]/div[2]/div[3]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/button[1]"))
        )
        edit_button.click()
        cp.settemplatename(self.templatename)
        cp.setlogoimg(self.absolute_path5)
        cp.clickonimgsave()
        cp.setprogramname(self.programname)
        cp.setcategoryname(self.categoryname)
        cp.setsign(self.absolute_path3)
        cp.clickonimgsave()
        cp.setsignname(self.signname)
        cp.setsigndesignation(self.signdesignation)
        cp.clickontemplatesave()

        xpath = "//div[contains(text(),'Created and unpublished successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_templatecreation.png")
            assert False
        time.sleep(3)
        cp.clickonedittemplate()
        time.sleep(1)

        # self.cp.scroll_to_end_of_page()
        cp.scrollnote()
        cp.clickontemplatepublish()

        xpath = "//div[contains(text(),'Updated and published successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_templateupdate.png")
            assert False
        time.sleep(3)
        cp.clickoncertification()
        cp.clickonmarkingsystemnew()
        cp.setcertificationname(certificationname)
        cp.setcertificationdescription(self.certificationdescription)
        cp.scrollupto()
        cp.clickonpublic()
        cp.clickonselectmarkingsystemfield()
        cp.clickonacronymselect()
        cp.clickonquestionnumber()
        cp.setquestionnumber(self.questionnumber)
        cp.setMarks(self.marks)
        cp.setreapply(self.reapply)
        cp.clickonminutes()
        cp.clickonminutestime()
        cp.clickonoptions()
        cp.clickonselectoption()
        cp.clickonselecttemplate()
        cp.clickontemplateclick()
        cp.clickoncertificatesave()
        cp.clickoncertificateconfirmsave()

        xpath = "//div[contains(text(),'Created and unpublished successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificationcreation.png")
            assert False
        time.sleep(3)
        cp.setsearch(certificationname)
        time.sleep(2)
        driver.find_element(By.XPATH, "//span[text()='" + certificationname + "']").click()
        cp.clickoncertificateedit()
        cp.clickoncertificatepublish()
        cp.clickoncertificateconfirmpublish()

        xpath = "//div[contains(text(),'Updated and published successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificateupdate.png")
            assert False

    @pytest.mark.sani
    @pytest.mark.regression
    @pytest.mark.run(order=73)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_certificationverificationforrelationcompany(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_003_Certification **********")
        lp = LoginPage(driver)
        lp.setUserName(self.usernames4)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login successful **********")
        cp = Certification(driver)
        cp.clickoncertificationprogramme()
        cp.clickonmyexams()
        workbook = load_workbook("TestData/LoginData.xlsx")

        # Access the active worksheet
        worksheet = workbook.active

        certificationname = worksheet["A21"].value

        workbook.close()

        cp.setsearch(certificationname)

        time.sleep(3)
        if "Your certification is typically displayed as a document stating that as a professional, you've been trained, educated and are prepared to meet a specific set of criteria for your role." in driver.page_source:
            self.logger.info("********** Certification verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Certification verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificateverificationforrelationcompany.png")
            assert False
        time.sleep(3)
        cp.clickongetcertificate()
        cp.clickonexamcheckbox()
        time.sleep(3)
        if "TAKE A TEST" in driver.page_source:
            self.logger.info("********** Exam verification test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Exam verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_examverificationforrelationcompany.png")
            assert False
        time.sleep(3)

    @pytest.mark.sanity
    @pytest.mark.regression
    @pytest.mark.run(order=74)
    # @pytest.mark.skip(reason="Skipping this test")
    def test_certificationverificationforremployee(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("************* Test_004_Certification **********")
        lp = LoginPage(driver)
        lp.setUserName(self.usernames5)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Login successful **********")
        cp = Certification(driver)
        cp.clickoncertificationprogramme()
        workbook = load_workbook("TestData/LoginData.xlsx")

        # Access the active worksheet
        worksheet = workbook.active

        certificationname = worksheet["A21"].value

        workbook.close()

        cp.setsearch(certificationname)
        time.sleep(3)
        if "Your certification is typically displayed as a document stating that as a professional, you've been trained, educated and are prepared to meet a specific set of criteria for your role." in driver.page_source:
            self.logger.info("********** certification verification test is passed *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** certification verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_certificateverificationforemployee.png")
            assert False
        time.sleep(3)
        cp.clickongetcertificate()
        cp.clickonexamcheckbox()
        time.sleep(3)
        if "TAKE A TEST" in driver.page_source:
            self.logger.info("********** Exam verification test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Exam verification test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_examverificationforemployee.png")
            assert False
        time.sleep(3)

    if __name__ == '__main__':
        unittest.main(verbosity=2)














