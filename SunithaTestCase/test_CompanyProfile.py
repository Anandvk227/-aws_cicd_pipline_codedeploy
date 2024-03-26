import os
import time

import pytest
from openpyxl.reader.excel import load_workbook
# import self
from selenium import webdriver
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from pageObjects.LoginPage import LoginPage
from sunithaPageObjects.CompanyProfile import LoginPage
from utilities.customLogger import LogGen
from pageObjects.randomGen import randomGen
from utilities.readProperties import ReadConfig
from sunithaPageObjects.MyProfile import MyprofilePage
from GenericLib.BaseClass import BaseClass

class TestCompanyProfile():
    baseURL = ReadConfig.getApplicationURL()
    # username = ReadConfig.getUseremail()
    password = ReadConfig.getPassword()
    relative_three = "Files/three.png"
    absolute_path3 = os.path.abspath(relative_three)
    relative_four = "Files/four.png"
    absolute_path4 = os.path.abspath(relative_four)
    relative_five = "Files/five.png"
    absolute_path5 = os.path.abspath(relative_five)
    relative_six = "Files/six.jpg"
    absolute_path6 = os.path.abspath(relative_six)
    relative_seven = "Files/seven.jpg"
    absolute_path7 = os.path.abspath(relative_seven)

    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value

    workbook.close()
    # Load the existing workbook
    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A2'] = username


    # Save the workbook
    wb.save("TestData/LoginData.xlsx")

    # BannerPath = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/3.jpg"
    # Upload_AgainBanner = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/5.jpg"
    # ProfilePath = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/4.jpg"
    # UpdateProfile_Again = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/7.jpg"

    Companyname = randomGen.random_addressInput() #"Econtent"
    orgname = randomGen.random_first_name()    #"Software and IT"
    industry = "educational services"
    website = "https://www.google.com/"
    companysummary = randomGen.random_overviwDescription()
    address = randomGen.random_addressInput()
    pincode = randomGen.random_pinCode()
    domainname = "mailcatch.com"
    contactperson = randomGen.random_first_name()
    url = "https://www.instagram.com/"

    # awards_path = "C:/Users/HP/PycharmProjects/AutomationFramework/imageFiles/11.jpg"
    awardTitle = randomGen.random_addressInput()

    logger = LogGen.loggen()



    @pytest.mark.regression
    @pytest.mark.run(order=36)
    # @pytest.mark.skip(reason="skipping this test")
    def test_BannerImage(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setpassword(self.password)
        lp.clickLogin()
        lp.clickCompanyProfile()
        self.logger.info("****** TC_01	Verify the Banner image/upload/save/edit/delete *****")
        lp.BannerImageClick(self.absolute_path3)
        lp.SaveBannerImage()

        xpath = "//div[contains(text(), 'Banner image uploaded successfully.')]"
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found: {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_BannerImage.png")
            assert False

        my = MyprofilePage(driver)
        time.sleep(1)
        my.clickClosetoaster()
        lp.EditBanner()
        lp.UploadBannerImageAgain(self.absolute_path4)
        lp.SaveBannerImage()
        xpath = "//div[contains(text(), 'Banner image uploaded successfully.')]"
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found: {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_BannerImage.png")
            assert False

        time.sleep(1)
        my.clickClosetoaster()
        lp.EditBanner()
        lp.BannerImageRemove()
        xpath = "//div[contains(text(), 'Banner image removed successfully')]"
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found: {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_BannerImage.png")
            assert False

    @pytest.mark.run(order=37)
    # @pytest.mark.test
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_ProfileImage(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.lp = LoginPage(driver)
        self.lp.setUserName(self.username)
        self.lp.setpassword(self.password)
        self.lp.clickLogin()
        self.lp.clickCompanyProfile()
        self.logger.info("****** TC_03	Verify the Profile image Upload/edit/save/delete *****")
        self.lp.ProfileImageClick(self.absolute_path5)
        self.lp.SaveProfileImage()

        xpath = "//div[contains(text(), 'Profile image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_ProfileImage.png")
            assert False

        self.lp.EditProfile()
        self.lp.UploadProfileImage(self.absolute_path6)
        self.lp.SaveProfileImage()

        xpath = "//div[contains(text(),'Profile image uploaded successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_ProfileImage.png")
            assert False

        self.lp.EditProfile()
        self.lp.ProfileImageRemove()
        xpath = "//div[contains(text(), 'Profile image removed successfully.')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_ProfileImage.png")
            assert False

    @pytest.mark.run(order=38)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_OfficialDetails(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setpassword(self.password)
        lp.clickLogin()
        lp.clickCompanyProfile()
        self.logger.info("****** TC_05	Verify the displaying all the Official details  *****")
        lp.ofc_details()
        # lp.setCompany(self.Companyname)
        lp.setOrganisation(self.orgname)
        lp.RemoveIndustry()
        lp.setIndustry(self.industry)
        lp.clickIndustryName()
        lp.clickCalendar()
        lp.clickonpreviousmonth()
        lp.clickDate()
        lp.setwebsite(self.website)
        lp.clicksave()

        xpath = "// div[contains(text(), 'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_OfficialDetails.png")
            assert False

    @pytest.mark.run(order=39)
    # @pytest.mark.test
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skipping this test")
    def test_OverView(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Login Test****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setpassword(self.password)
        lp.clickLogin()
        # lp.clickNewsfeedModule()
        lp.clickCompanyProfile()
        lp.scrollIntoOverView()

        self.logger.info("****** TC_11	Verify the Overview, by Add, Save and Update  *****")
        time.sleep(2)
        lp.clickEdit()
        lp.setCompanySummary(self.companysummary)
        lp.setAddress(self.address)
        actions = ActionChains(driver)
        actions.send_keys(Keys.PAGE_DOWN).perform()
        driver.execute_script("window.scrollBy(0, 500);")
        time.sleep(2)
        lp.countrydropdown()
        lp.clickonselectindia()
        lp.clickState()
        lp.clickstatelistbox()
        lp.selectState()
        lp.clickcity()
        lp.selectCity()
        lp.setpincode(self.pincode)
        lp.setdomainName(self.domainname)
        lp.setcontactPerson(self.contactperson)
        lp.clicksave()
        xpath = "// div[contains(text(), 'Company profile updated')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_OverView.png")
            assert False

    @pytest.mark.run(order=40)
    @pytest.mark.test
    @pytest.mark.sunitha
    # @pytest.mark.skip(reason="skipping this test")
    def test_Awards(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Awards Test****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setpassword(self.password)
        lp.clickLogin()
        lp.clickCompanyProfile()
        lp.scrollIntoAwards()
        self.logger.info("****** TC_13	Verify the Awards, by Add, Save and Edit *****")
        lp.AwardsEdit()
        lp.ClickPreview(self.absolute_path7)
        lp.TitleInput(self.awardTitle)
        lp.SaveAward()
        xpath = "//div[contains(text(), 'Award added successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_Awards.png")
            assert False

    @pytest.mark.run(order=41)
    @pytest.mark.flaky(rerun=3, rerun_delay=2)
    @pytest.mark.regression
    @pytest.mark.test
    # @pytest.mark.skip(reason="skipping this test")
    def test_socialMediaLinks(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setpassword(self.password)
        lp.clickLogin()
        lp.clickCompanyProfile()
        self.logger.info("****** TC_16	Verify Social Media Links details, by Add, Save and Update  *****")
        lp.clicksocialMediaLinks()
        lp.clicksocilmedialist()
        lp.clickSocialMediaName()
        lp.setUrl(self.url)
        lp.clickUrlsave()
        xpath = "//div[contains(text(),'Profile updated  successfully')]"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_socialMediaLinks.png")
            assert False

    @pytest.mark.run(order=42)
    @pytest.mark.falky(rerun=3, rerun_delay=2)

    @pytest.mark.regression
    def test_ClickingElements(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Clicking Elements Test****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setpassword(self.password)
        lp.clickLogin()

        lp.clickCompanyProfile()
        time.sleep(1)
        self.logger.info("****** TC_19	Verify Networks by click on that  *****")
        lp.NetworksClick()
        xpath = "//button[normalize-space()='MY NETWORKS']"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_ClickingElements.png")
            assert False

        driver.back()
        time.sleep(3)
        self.logger.info("****** TC_20	verify the certificate by click on that  *****")
        lp.CertificationsClick()
        xpath = "//button[normalize-space()='Create certification']"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_ClickingElements.png")
            assert False

        driver.back()
        time.sleep(3)
        self.logger.info("****** TC_18	Verify Resources by click on that *****")
        lp.ResourceClick()
        xpath = "//button[normalize-space()='My RESOURCES']"
        try:
            # Use WebDriverWait to wait for the element to be present
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot(".\\Screenshots\\" + "test_ClickingElements.png")
            assert False

        driver.back()
        time.sleep(2)
        self.logger.info("****** TC_25	verify the setting tab by clicking on it *****")
        lp.Settings()
        self.logger.info("****** TC_24	verify the news feed by clicking on that  *****")
        lp.NewsFeed()

    if __name__ == '__main__':
        unittest.main(verbosity=2)


