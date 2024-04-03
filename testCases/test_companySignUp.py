import time
import unittest
import pytest
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from selenium.common import NoSuchElementException, StaleElementReferenceException, TimeoutException
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pageObjects.companyListingPage import companyListingPage
from pageObjects.LoginPage import LoginPage
from testCases.mailboxCode import mailboxCode
# from testCases.conftest import setup
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from pageObjects.companySignUpPage import companySignUpPage
from pageObjects.randomGen import randomGen
from selenium import webdriver
from GenericLib.BaseClass import BaseClass
from openpyxl import workbook
import re


class TestSignUp():
    baseURL = ReadConfig.getApplicationURL()
    setSearchIndustryType = "Information Technology"
    password = ReadConfig.getPassword()

    email = randomGen.random_email()
    first_name = randomGen.random_first_name()
    company_name = randomGen.random_company_name()
    phone_number = randomGen.random_phone_number()
    # Load the existing workbook
    workbook = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    worksheet = workbook.active

    # Update the existing cells with new data
    worksheet['A2'] = email
    worksheet['B2'] = first_name
    worksheet['C2'] = company_name
    worksheet['D2'] = phone_number

    # Save the workbook
    workbook.save("TestData/LoginData.xlsx")
    workbook.close()
    time.sleep(5)
    logger = LogGen.loggen()
    @pytest.mark.run(order=1)
    @pytest.mark.regression
    # @pytest.mark.test
    # @pytest.mark.flaky(rerun=3, rerun_delay=2)
    def test_SignUpwithValid(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("********TS_1	TC1_1	Verify the Signup functionality. with positive data. ***********")
        self.logger.info("******** User is on Login page ***********")
        sp = companySignUpPage(driver)
        sp.clicksignuplink()
        sp.clickCompanysignupButton()
        self.logger.info("******** user is in company signup page ***********")
        self.logger.info("******** Entering valid data into the fields ***********")
        sp.setCompanyName(self.company_name)
        sp.setSearchIndustryType(self.setSearchIndustryType)
        sp.selectCompany()
        sp.setContactName(self.first_name)
        sp.setEmail(self.email)
        sp.clickcountrydd()
        sp.clickindia()
        workbook = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        worksheet = workbook.active
        email = worksheet["A2"].value

        sp.clickstatedd()
        sp.clickTelangana()
        sp.clickcitydd()
        sp.clickHyderabad()
        sp.setPhone(self.phone_number)
        sp.setPassword(self.password)
        sp.setConfirmPassword(self.password)
        sp.clicktermsConditions()
        time.sleep(2)
        self.logger.info("******** Clicking on signup button ***********")
        self.logger.info("******** TC1_2  Verify that a User can Successfully Sign Up with OTP  ***********")
        sp.clicksignupNow()
        time.sleep(2)

        # Execute JavaScript to open a new tab
        driver.execute_script("window.open('about:blank', '_blank');")

        # Perform actions in the new tab (if needed)
        # For example:
        driver.switch_to.window(driver.window_handles[1])
        self.logger.info("******** Opening new url in another tab for Email OTP ***********")
        time.sleep(1)
        driver.get("http://mailcatch.com/en/disposable-email")
        time.sleep(1)
        yopmail = driver.find_element(By.XPATH, "//input[@name='box']")
        yopmail.send_keys(email + Keys.ENTER)
        time.sleep(1)

        reload_button = driver.find_element(By.XPATH, "//img[@title='Reload']")

        # Click the Reload button every second until the subject is displayed or a maximum time is reached
        max_wait_time = 60  # Set your maximum wait time in seconds
        start_time = time.time()

        while time.time() - start_time < max_wait_time:
            reload_button.click()

            try:
                # Check if the subject is displayed
                subject = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//td[@class='subject']"))
                )
                subject.click()
                break  # Break out of the loop if subject is displayed
            except StaleElementReferenceException:
                print("StaleElementReferenceException occurred. Retrying...")
                continue  # Retry the loop if StaleElementReferenceException occurs
            except TimeoutException:
                time.sleep(1)

        iframeElement = driver.find_element(By.ID, "emailframe")
        driver.switch_to.frame(iframeElement)

        # Code outside the loop will be executed after the loop or when a TimeoutException occurs
        otp = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//body"))
        )
        driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        # This code is for QA ENV
        otp = driver.find_element(By.XPATH, "//body")
        driver.execute_script("arguments[0].scrollIntoView(true);", otp)
        time.sleep(0.5)

        confirmation_code = otp.text
        getOTP = re.search(r'\b\d+\b', confirmation_code).group()
        print(getOTP)

        self.logger.info("******** Switching back and entering the otp ***********")
        driver.switch_to.default_content()

        driver.switch_to.window(driver.window_handles[0])

        sp.setOtp(getOTP)

        time.sleep(2)
        self.logger.info(
            "******** TC3_1 Verify the Signup page OTP page,  Verify that a user can successfully verify their account with a valid OTP. ***********")
        sp.clickVerifyButton()
        sp.clickContinueToLogin()
        self.logger.info("******** Company Sign Up successful ***********")
        self.logger.info("******** Entering the sig up credentials for Login ***********")

        # Read data from specific cells
        email = worksheet['A2'].value
        lp = LoginPage(driver)
        lp.setUserName(email)
        lp.setPassword(self.password)
        lp.clickLogin()
        lp.clickcreatePost()
        self.logger.info("******** Login successful ***********")
        act_Text = lp.newsFeedText()

        if act_Text == "Create News Feed":
            assert True
            self.logger.info("********* SignUp Test is Passed ***********")

        else:
            driver.save_screenshot(".\\Screenshots\\" + "test_SignUpwithValid.png")
            self.logger.error("********* SignUp Test is Failed ***********")
            driver.close()
            assert False

        time.sleep(3)
        driver.find_element(By.XPATH, "//div[@class='flexAutoRow alignCntr pdngHXS']").click()
        # driver.close()

    @pytest.mark.run(order=2)
    @pytest.mark.regression
    @pytest.mark.skip
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_ListingSignUpCompany(self, driver):
        # Url = "https://preprodanalytics.inlynk.com/license"
        # username = "sowjanyapreprod@yopmail.com"
        # password = "Inlink@123"
        Url = "https://testanalytics.inlynk.com/"
        username = "ntpc@yopmail.com"
        password = "Inlink@123"
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(Url)
        lp = LoginPage(driver)
        clp = companyListingPage(driver)
        clp.setUserName(username)
        clp.setPassword(password)
        lp.clickLogin()

        clp.clickLicense()
        workbook = load_workbook("TestData/LoginData.xlsx")

        # Select the active worksheet
        worksheet = workbook.active
        company = worksheet['C2'].value
        clp.setsearchFiled(company)
        xpath = "//span[contains(text(),'" + company + "')]"
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_ListingSignUpCompany.png")
            driver.close()
            driver.quit()
            assert False
        time.sleep(1)
        element.click()
        clp.clicksubscription()
        clp.clickedit()
        time.sleep(3)
        clp.clicklisted()
        clp.clickupdate()
        xpath = clp.ListedToast
        # Use WebDriverWait to wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_ListingSignUpCompany.png")
            driver.close()
            driver.quit()
            assert False

    if __name__ == '__main__':
        pytest.main(['-v', '-p', 'pytest_ordering', __file__])
