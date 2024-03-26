import time
import unittest

import pytest
from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.common import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pageObjects.networksPage import networksPage

from pageObjects.LoginPage import LoginPage
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen

from GenericLib.BaseClass import BaseClass

class TestNetworks():
    baseURL = ReadConfig.getApplicationURL()
    workbook = load_workbook("TestData/LoginData.xlsx")

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["A2"].value
    RMname = worksheet["J2"].value
    RMname2 = worksheet["J3"].value
    RMname3 = worksheet["J4"].value
    password = ReadConfig.getPassword()
    CompanyManufacture = worksheet["H2"].value
    EmailManufacture = worksheet["I2"].value
    CompanyPartner = worksheet["H3"].value
    EmailPartner = worksheet["I3"].value
    CompanyDistributor = worksheet["H4"].value
    EmailDistributor = worksheet["I4"].value
    CompanyVendor = worksheet["H5"].value
    EmailVendor = worksheet["I5"].value

    workbook.close()

    logger = LogGen.loggen()


    @pytest.mark.run(order=22)
    # @pytest.mark.regression
    # @pytest.mark.skip
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_RejectConnectionCompanyAsManufacturer(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Network Connection Test****")
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value

        lp = LoginPage(driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + CompEmail + " and password " + self.password)
        lp.setUserName(CompEmail)
        lp.setPassword(self.password)

        lp.clickLogin()
        np = networksPage(driver)
        np.clickNetworks()
        np.setsearchField(companyName)
        time.sleep(3)
        # Verifying and Clicking on the company
        # with an explicit wait
        xpath = "//span[contains(text(),'" + companyName + "')]"
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        element.click()
        np.clickConnectButton()
        time.sleep(2)
        np.clickDropDownList()
        np.clickManufacturer()
        np.setRM_searchField(self.RMname)
        np.clickSelectRM()
        np.clickcheckbox()
        np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False

        np.clickOKButton()
        lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        np.clickNetworks()
        np.clickPendingTab()
        np.setsearchField(self.CompanyManufacture)
        # Assuming `driver` is your WebDriver instance
        company_manufacture_xpath = "//span[contains(text(),'" + self.CompanyManufacture + "')]"

        # Explicit wait for the element to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, company_manufacture_xpath))
        )

        # Click the element
        element.click()
        np.clickRejectButton()
        np.setTextarea(self.CompanyManufacture)
        np.clickReject2Button()
        time.sleep(4)
        act_Text = np.Text_Connection_rejected_successfully()
        if act_Text == "Connection rejected successfully":
            assert True
            self.logger.info("********* Company Connection rejected Successfully ***********")
            # driver.close()
        else:
            driver.save_screenshot(".\\Screenshots\\" + "connection_reject_fail.png")
            self.logger.error("********* Company reject fail ***********")
            # driver.close()
            assert False

    @pytest.mark.run(order=25)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_ConnectionCompanyAsManufacturer(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Network Connection Test****")
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value
        lp = LoginPage(driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + CompEmail + " and password " + self.password)
        lp.setUserName(CompEmail)
        lp.setPassword(self.password)

        lp.clickLogin()
        np = networksPage(driver)
        np.clickNetworks()
        np.setsearchField(companyName)
        # Verifying and Clicking on the company
        xpath = "//span[contains(text(),'" + companyName + "')]"
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        element.click()
        np.clickConnectButton()
        time.sleep(2)
        np.clickDropDownList()
        np.clickManufacturer()
        np.setRM_searchField(self.RMname)
        np.clickSelectRM()
        np.clickcheckbox()
        np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False

        np.clickOKButton()
        lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        np.clickNetworks()
        np.clickPendingTab()
        np.setsearchField(self.CompanyManufacture)
        driver.find_element(By.XPATH, "//span[contains(text(),'" + self.CompanyManufacture + "')]").click()
        np.clickApproveButton()
        time.sleep(2)
        np.clickAcceptButton()
        time.sleep(2)
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsManufacturer.png")
                assert False

    @pytest.mark.run(order=26)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_ConnectionCompanyAsPartner(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Network Connection Test****")
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmailpartner = ws['I3'].value
        companyName = ws['C2'].value
        lp = LoginPage(driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + CompEmailpartner + " and password " + self.password)
        lp.setUserName(CompEmailpartner)
        lp.setPassword(self.password)
        lp.clickLogin()
        np = networksPage(driver)
        np.clickNetworks()
        np.setsearchField(companyName)
        # Verifying and Clicking on the company
        # with an explicit wait
        xpath = "//span[contains(text(),'" + companyName + "')]"
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        element.click()
        np.clickConnectButton()
        time.sleep(2)
        np.clickDropDownList()
        time.sleep(2)
        np.clickpartner()
        np.setRM_searchField(self.RMname2)
        np.clickSelectRM()
        np.clickcheckbox()
        np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsPartner.png")
                assert False

        np.clickOKButton()
        lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + self.username + " and Password:" + self.password)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        np.clickNetworks()
        np.clickPendingTab()
        np.setsearchField(self.CompanyPartner)
        # Assuming `driver` is your WebDriver instance
        company_partner_xpath = "//span[contains(text(),'" + self.CompanyPartner + "')]"

        # Explicit wait for the element to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, company_partner_xpath))
        )

        # Click the element
        element.click()
        np.clickApproveButton()
        time.sleep(2)
        np.clickAcceptButton()
        time.sleep(2)
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.info(f"Error text not found within the specified time: {text}")
                driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsPartner.png")
                assert False

    @pytest.mark.run(order=27)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_ConnectionCompanyAsShareHolder(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Network Connection Test****")
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmailshareholder = ws['I4'].value
        companyName = ws['C2'].value
        lp = LoginPage(driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + CompEmailshareholder + " and password " + self.password)
        lp.setUserName(CompEmailshareholder)
        lp.setPassword(self.password)
        lp.clickLogin()
        np = networksPage(driver)
        np.clickNetworks()
        np.setsearchField(companyName)
        # Verifying and Clicking on the company
        # with an explicit wait
        xpath = "//span[contains(text(),'" + companyName + "')]"
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        element.click()
        np.clickConnectButton()
        time.sleep(2)
        np.clickDropDownList()
        time.sleep(1)
        np.clickshareHolder()
        np.setRM_searchField(self.RMname3)
        np.clickSelectRM()
        np.clickcheckbox()
        np.clickButtonConnect2()
        # List of error texts to validate
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing error texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                error_elements = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if error_elements:
                    self.logger.info(f"Found error text: {text}")
                    assert True

            except TimeoutException:
                self.logger.error(f"Error text not found within the specified time: {text}")
                driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsShareHolder.png")
                assert False

        np.clickOKButton()
        lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        lp = LoginPage(driver)
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + self.EmailDistributor + " and Password:" + self.password)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        np.clickNetworks()
        np.clickPendingTab()
        np.setsearchField(self.CompanyDistributor)
        # Assuming `driver` is your WebDriver instance
        company_distributor_xpath = "//span[contains(text(),'" + self.CompanyDistributor + "')]"

        # Explicit wait for the element to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, company_distributor_xpath))
        )

        # Click the element
        element.click()
        np.clickApproveButton()
        time.sleep(4)
        np.clickAcceptButton()

        # Validate success message after approval
        Verify_texts = [
            "Network Connections",
        ]

        # Find elements containing success texts and validate
        for text in Verify_texts:
            try:
                # Wait up to 10 seconds until at least one element with the specified text is present
                success_elements = WebDriverWait(driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH, f"//*[contains(text(), '{text}')]"))
                )

                if success_elements:
                    self.logger.info(f"Found success text: {text}")
                    assert True

            except TimeoutException:
                self.logger.error(f"Success text not found within the specified time: {text}")
                driver.save_screenshot(".\\Screenshots\\" + "test_ConnectionCompanyAsShareHolder.png")
                assert False

    # @pytest.mark.smoke(order=3)
    @pytest.mark.run(order=23)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    # @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_follow_and_unfollow_company(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Network Connection Test****")
        self.logger.info("****TC_05	Verify Follow tab****")
        lp = LoginPage(driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + self.username + " and password " + self.password)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        np = networksPage(driver)
        np.clickNetworks()
        np.setsearchField(self.CompanyManufacture)
        # Verifying and Clicking on the company
        company_manufacture_xpath = "//span[contains(text(),'" + self.CompanyManufacture + "')]"

        # Explicit wait for the element to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, company_manufacture_xpath))
        )

        # Click the element
        element.click()
        time.sleep(1)
        np.clickFollowButton()
        time.sleep(1)
        driver.back()
        np.clickFOLLOWTab()

        np.setsearchField(self.CompanyManufacture)

        if "Following" in driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_follow_and_unfollow_company.png")
            assert False
        lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value
        CompanyManufacture = ws['H2'].value
        lp = LoginPage(driver)
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        lp.setUserName(CompEmail)
        lp.setPassword(self.password)
        lp.clickLogin()
        np.clickNetworks()
        np.clickFOLLOWTab()
        time.sleep(1)
        np.setsearchField(companyName)
        xpath = "//span[contains(text(),'" + companyName + "')]"
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        element.click()
        time.sleep(3)
        if "Follow" in driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_follow_and_unfollow_company.png")
            assert False
        lp.clickLogout()
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + self.username + " and password " + self.password)
        lp.setUserName(self.username)
        lp.setPassword(self.password)

        lp.clickLogin()
        np = networksPage(driver)
        np.clickNetworks()
        np.clickFOLLOWTab()
        np.setsearchField(self.CompanyManufacture)
        # Verifying and Clicking on the company
        company_manufacture_xpath = "//span[contains(text(),'" + self.CompanyManufacture + "')]"

        # Explicit wait for the element to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, company_manufacture_xpath))
        )

        # Click the element
        element.click()
        time.sleep(4)
        if "Following" in driver.page_source:
            self.logger.info("********** Successfully see the page *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** following company is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_follow_and_unfollow_company.png")
            assert False
        time.sleep(2)
        np.clickFollowingButton()
        np.clickUnfollowButton()
        time.sleep(2)

    @pytest.mark.run(order=24)
    @pytest.mark.regression
    # @pytest.mark.skip(reason="skip for now")
    def test_block_and_unblock_the_followed_company(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Network Connection Test****")
        self.logger.info("****TC_05	Verify Follow tab****")
        lp = LoginPage(driver)
        self.logger.info(
            "Entering SuperAdmin Credentials for login username " + self.username + " and password " + self.password)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        np = networksPage(driver)
        np.clickNetworks()
        np.setsearchField(self.CompanyManufacture)
        # Verifying and Clicking on the company
        company_manufacture_xpath = "//span[contains(text(),'" + self.CompanyManufacture + "')]"

        # Explicit wait for the element to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, company_manufacture_xpath))
        )

        # Click the element
        element.click()
        time.sleep(2)
        np.clickFollowButton()
        time.sleep(1)
        driver.back()
        np.clickFOLLOWTab()
        time.sleep(2)
        np.setsearchField(self.CompanyManufacture)
        company_manufacture_xpath = "//span[contains(text(),'" + self.CompanyManufacture + "')]"

        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, company_manufacture_xpath))
        )

        element.click()
        time.sleep(2)
        if "Follow" in driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            self.logger.error("************** following company is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_block_and_unblock_the_followed_company.png")
            assert False

        lp.clickLogout()
        self.logger.info("******** Entering Requested company credentials for Network Connection ***********")
        # Read data from specific cells
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        CompEmail = ws['I2'].value
        companyName = ws['C2'].value
        CompanyManufacture = ws['H2'].value
        lp = LoginPage(driver)
        self.logger.info(
            "Entering Requested company Credentials for Approve request Username:" + CompEmail + " and Password:" + self.password)
        lp.setUserName(CompEmail)
        lp.setPassword(self.password)
        lp.clickLogin()
        np.clickNetworks()
        np.clickFOLLOWTab()
        time.sleep(1)
        np.setsearchField(companyName)
        xpath = "//span[contains(text(),'" + companyName + "')]"
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        element.click()
        time.sleep(1)
        driver.back()
        np.ClickBlockCompany()
        np.ClickConfirmBlock()

        xpath = "//div[contains(text(), 'You have blocked " + companyName + "')]"
        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, xpath))
            )
            self.logger.info(f"Text Found : {element.text}")
            assert True
        except:
            self.logger.info(f"Text Not Found")
            driver.save_screenshot("\\Screenshots\\" + "test_block_and_unblock_the_followed_company.png")
            assert False

        np.clickBlocklistTab()
        time.sleep(1)
        np.setsearchField(companyName)
        xpath = "//span[contains(text(),'" + companyName + "')]"
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, xpath))
        )
        element.click()
        time.sleep(3)
        np.ClickUnblockCompany()
        np.ClickConfirmUnblock()
        time.sleep(3)
        if "You have unblocked " + companyName + "" in driver.page_source:
            self.logger.info("********** Following company successfully *********")

        else:
            self.logger.error("************** following company is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_block_and_unblock_the_followed_company.png")
            assert False

if __name__ == "__main__":
    unittest.main()

