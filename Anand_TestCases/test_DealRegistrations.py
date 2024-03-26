import time

import pytest

from openpyxl.reader.excel import load_workbook

from pageObjects.LoginPage import LoginPage
from utilities.customLogger import LogGen
from utilities.readProperties import ReadConfig
from Anand_PageObjects.DealRegistration import dealregistration
from pageObjects.randomGen import randomGen
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from GenericLib.BaseClass import BaseClass



class Test_Create_DealwithNetworkCompany():
    baseURL = ReadConfig.getApplicationURL()
    workbook = load_workbook("TestData/LoginData.xlsx")
    deptname = randomGen.random_first_name()
    comapanyname="j industries"
    nameofcompany="Instavc Technologies"
    address="Q3-A4,Cyber Towers,Hi tech city"
    department="Outsidesales "+deptname
    departmenttwo="BPo"+deptname
    departmentthree="InsideSales"+deptname
    departmentfour="internationalMarketing"+deptname
    departmentfive="DealNetworks"+deptname
    departmentsix="ITcompany"+deptname

    wb = load_workbook("TestData/LoginData.xlsx")

    # Select the active worksheet
    ws = wb.active

    # Update the existing cells with new data
    ws['A16'] = department
    ws['C16'] = departmentfour



    # Save the workbook
    wb.save("TestData/LoginData.xlsx")
    contname="Anand"
    contemail="anand.n@instavc.com"
    contnumber="8363453672"
    nameofmanager="a"
    oppdetails="CPU & Laptops"
    currencydetails="INR"
    valueofdeal="92826"
    reason="Present deal items are not available"
    accountmanager="persona"
    search="j industries"
    searchtwo="mahindra"

    # Access the active worksheet
    worksheet = workbook.active

    username = worksheet["B14"].value
    username1=worksheet["B15"].value
    username2=worksheet["B16"].value
    password = ReadConfig.getPassword()

    workbook.close()

    logger=LogGen.loggen()

    @pytest.mark.regression
    @pytest.mark.test
    @pytest.mark.run(order=75)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    # @pytest.mark.skip(reason="skipping this Test")
    def test_deal_Create_Approve_Relation_Company(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****TC_1 Verify the OEM Company Create The Deal and Approve by partner company****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        if "News Feed" in driver.page_source:
            self.logger.info("********** OEM Company Login successfully *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Login failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "OEM_login_fail.png")
            assert False

        deal = dealregistration(driver)
        deal.clickdealtab()
        if "Deal registration" in driver.page_source:
            self.logger.info(
                "********** My Active Deals page & New Button will display for Creating new deal *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Deal Registration Page is not open successfully **********")
            driver.save_screenshot(".\\Screenshots\\" + "deal_registration_fail_to_open.png")
            assert False

        deal.clickonnewdeal()
        deal.networkcomapnyselect(self.comapanyname)
        deal.companydetails(self.nameofcompany)
        deal.adreesdetails(self.address)
        deal.countrydropdown()
        deal.selectcountry()
        deal.statedropdown()
        deal.selectstate()
        time.sleep(1)
        deal.citydropdown()
        deal.selectcity()
        deal.departmentname(self.departmenttwo)
        deal.contactname(self.contname)
        deal.contactemail(self.contemail)
        deal.selectenterprise()
        deal.accountmanager(self.nameofmanager)
        deal.relationmanager()
        deal.enteropportunity(self.oppdetails)
        deal.dealvalue(self.valueofdeal)
        deal.savedeal()
        if "Deal Opportunity Created" in driver.page_source:
            self.logger.info("********** Deal Created Successfully *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Deal not Created Successfully **********")
            driver.save_screenshot(".\\Screenshots\\" + "deal_creation_fail.png")
            assert False

        deal.okaybutton()
        time.sleep(1)
        lp.clickLogout()
        if "Login" in driver.page_source:
            self.logger.info("********** OEM Company Logout successfully *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Logout failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "OEM_logout_fail.png")
            assert False

        lp.setUserName(self.username1)
        lp.setPassword(self.password)
        lp.clickLogin()
        if "News Feed" in driver.page_source:
            self.logger.info("********** partner Company Login successfully *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** partner Company Login failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "partner_login_fail.png")
            assert False

        deal.clickdealtab()
        deal.networkdeals()
        if "Deal registration" in driver.page_source:
            self.logger.info("********** Network Deals page display  *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Network deals page not open **********")
            driver.save_screenshot(".\\Screenshots\\" + "network_deals_fail_to_open.png")
            assert False

        deal.pendingdeals()
        deal.searchmydealcompany(self.searchtwo)
        time.sleep(1)
        deal.dealcompany()
        if "Instavc Technologies" in driver.page_source:
            self.logger.info("********** Latest created deal is Displaying *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Not Displaying Latest Created Deal **********")
            driver.save_screenshot(".\\Screenshots\\" + "failed_to_open_selected_company_pending_deals.png")
            assert False

        deal.selectnewdeal()
        time.sleep(1)
        if "Deal Opportunity Details" in driver.page_source:
            self.logger.info("********** Deal created and opened successfully by that created Deal *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Failed to open created deal **********")
            driver.save_screenshot(".\\Screenshots\\" + "deal_open_fail.png")
            assert False

        deal.statusdropdown()
        deal.approved()
        deal.confirmtoapprove()
        time.sleep(2)
        if "Deal Approved Successfully" in driver.page_source:
            self.logger.info("********** test_deal_Create_Approved_Relation_Company successfully *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Approved deal failed  **********")
            driver.save_screenshot(".\\Screenshots\\" + "approved_deal_fail.png")
            assert False

    # @pytest.mark.skip(reason="skipping this Test")

    @pytest.mark.regression
    @pytest.mark.run(order=76)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_deal_Create_Reject_Relation_Company(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****TC_02 Verify the OEM Company Create The Deal and reject the deal by partner company****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        if "News Feed" in driver.page_source:
            self.logger.info("********** OEM Company Login successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Login failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "OEM_login_fail.png")
            assert False

        deal = dealregistration(driver)
        deal.clickdealtab()
        if "Deal registration" in driver.page_source:
            self.logger.info(
                "********** My Active Deals page & New Button will display for Creating new deal *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal Registration Page is not open successfully **********")
            driver.save_screenshot(".\\Screenshots\\" + "deal_registration_fail_to_open.png")
            assert False
        deal.clickonnewdeal()
        deal.networkcomapnyselect(self.comapanyname)
        deal.companydetails(self.nameofcompany)
        deal.adreesdetails(self.address)
        deal.countrydropdown()
        deal.selectcountry()
        deal.statedropdown()
        deal.selectstate()
        time.sleep(1)
        deal.citydropdown()
        deal.selectcity()
        deal.departmentname(self.departmentthree)
        deal.contactname(self.contname)
        deal.contactemail(self.contemail)
        # deal.contactnumber(self.contnumber)
        deal.selectenterprise()
        deal.accountmanager(self.nameofmanager)
        deal.relationmanager()
        deal.enteropportunity(self.oppdetails)
        # deal.currency(self.currencydetails)
        deal.dealvalue(self.valueofdeal)
        deal.savedeal()
        if "Deal Opportunity Created" in driver.page_source:
            self.logger.info("********** Deal Created Successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal not Created Successfully **********")
            driver.save_screenshot(".\\Screenshots\\" + "deal_creation_fail.png")
            assert False

        deal.okaybutton()
        lp.clickLogout()
        if "Login" in driver.page_source:
            self.logger.info("********** OEM Company Logout successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Logout failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "OEM_logout_fail.png")
            assert False
        lp.setUserName(self.username1)
        lp.setPassword(self.password)
        lp.clickLogin()
        if "News Feed" in driver.page_source:
            self.logger.info("********** partner Company Login successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** partner Company Login failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "partner_login_fail.png")
            assert False
        deal.clickdealtab()
        deal.networkdeals()
        if "Deal registration" in driver.page_source:
            self.logger.info("********** Network Deals page display  *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Network deals page not open **********")
            driver.save_screenshot(".\\Screenshots\\" + "network_deals_fail_to_open.png")
            assert False
        deal.pendingdeals()
        deal.searchmydealcompany(self.searchtwo)
        time.sleep(1)
        deal.dealcompany()
        if "Instavc Technologies" in driver.page_source:
            self.logger.info("********** Latest created deal is Displaying *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Not Displaying Latest Created Deal **********")
            driver.save_screenshot(".\\Screenshots\\" + "failed_to_open_selected_company_pending_deals.png")
            assert False
        deal.selectnewdeal()
        time.sleep(1)
        if "Deal Opportunity Details" in driver.page_source:
            self.logger.info("********** Deal created and opened successfully by that created Deal *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Failed to open created deal **********")
            driver.save_screenshot(".\\Screenshots\\" + "deal_open_fail.png")
            assert False
        deal.statusdropdown()
        deal.clickreject()
        deal.reasontoreject(self.reason)
        deal.clickonreject()
        time.sleep(2)
        if "Deal Rejected Successfully" in driver.page_source:
            self.logger.info("********** test_deal_Create_rejected_Relation_Company successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Rejected deal failed  **********")
            driver.save_screenshot(".\\Screenshots\\" + "rejected_deal_fail.png")
            assert False

    # @pytest.mark.skip(reason="skipping this Test")

    @pytest.mark.regression
    @pytest.mark.run(order=77)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_deal_Create_Edit_Approve_Through_Notification(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****TC_03 Create a Deal with OEM Company and Approve the Deal through Notification ****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        if "News Feed" in driver.page_source:
            self.logger.info("********** OEM Company Login successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Login failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "OEM_login_fail.png")
            assert False

        deal = dealregistration(driver)
        deal.clickdealtab()
        if "Deal registration" in driver.page_source:
            self.logger.info(
                "********** My Active Deals page & New Button will display for Creating new deal *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal Registration Page is not open successfully **********")
            driver.save_screenshot(".\\Screenshots\\" + "deal_registration_fail_to_open.png")
            assert False
        deal.clickonnewdeal()
        deal.networkcomapnyselect(self.comapanyname)
        deal.companydetails(self.nameofcompany)
        deal.adreesdetails(self.address)
        deal.countrydropdown()
        deal.selectcountry()
        deal.statedropdown()
        deal.selectstate()
        time.sleep(1)
        deal.citydropdown()
        deal.selectcity()
        deal.departmentname(self.department)
        deal.contactname(self.contname)
        deal.contactemail(self.contemail)
        # deal.contactnumber(self.contnumber)
        deal.selectenterprise()
        deal.accountmanager(self.nameofmanager)
        deal.relationmanager()
        deal.enteropportunity(self.oppdetails)
        # deal.currency(self.currencydetails)
        deal.dealvalue(self.valueofdeal)
        deal.savedeal()
        if "Deal Opportunity Created" in driver.page_source:
            self.logger.info("********** Deal Created Successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** Deal not Created Successfully **********")
            driver.save_screenshot(".\\Screenshots\\" + "deal_creation_fail.png")
            assert False
        deal.okaybutton()
        lp.clickLogout()
        if "Login" in driver.page_source:
            self.logger.info("********** OEM Company Logout successfully *********")

        else:
            # Log and take a screenshot
            self.logger.error("************** OEM Company Logout failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "OEM_logout_fail.png")
            assert False
        lp.setUserName(self.username1)
        lp.setPassword(self.password)
        lp.clickLogin()

        deal.clickonnotification()

        # changes excel sheet data
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        department = ws['A16'].value
        element = driver.find_element(By.XPATH,
                                      "//span[text()='Anand Mahindra created a new deal with Instavc Technologies in Department " + department + "']")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_deal_Create_Edit_Approve_Through_Notification.png")
            assert False

        element.click()
        deal.clickhere()
        deal.dealdetailspage()
        deal.approveediteddeal()
        deal.confirmupdatedeal()
        deal.closeapprovedtab()
        time.sleep(1)

    # @pytest.mark.skip(reason="skipping this Test")

    @pytest.mark.regression
    @pytest.mark.run(order=78)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_deal_Reject_Through_Notification(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Login Test****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal = dealregistration(driver)
        deal.clickdealtab()
        deal.clickonnewdeal()
        deal.networkcomapnyselect(self.comapanyname)
        deal.companydetails(self.nameofcompany)
        deal.adreesdetails(self.address)
        deal.countrydropdown()
        deal.selectcountry()
        deal.statedropdown()
        deal.selectstate()
        time.sleep(1)
        deal.citydropdown()
        deal.selectcity()
        time.sleep(2)
        deal.departmentname(self.departmentfour)
        time.sleep(2)
        deal.contactname(self.contname)
        deal.contactemail(self.contemail)
        # deal.contactnumber(self.contnumber)
        deal.selectenterprise()
        time.sleep(2)
        deal.accountmanager(self.nameofmanager)
        deal.relationmanager()
        deal.enteropportunity(self.oppdetails)
        # deal.currency(self.currencydetails)
        deal.dealvalue(self.valueofdeal)
        time.sleep(1)
        deal.savedeal()
        deal.okaybutton()
        time.sleep(2)
        lp.clickLogout()
        time.sleep(1)
        lp.setUserName(self.username1)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal.clickonnotification()

        # changes excel sheet data
        wb = load_workbook("TestData/LoginData.xlsx")
        ws = wb.active
        departmentfour = ws['C16'].value
        element = driver.find_element(By.XPATH,
                                      "//span[text()='Anand Mahindra created a new deal with Instavc Technologies in Department " + departmentfour + "']")
        # assert element.text == first_name, f"Expected '{first_name}' but found '{element.text}'"

        if element:
            self.logger.info(f"Found Employee name : {element.text}")
            assert True
            # driver.quit()
        else:
            self.logger.info(f"Employee name not found: {element.text}")
            driver.save_screenshot(".\\Screenshots\\" + "test_deal_Reject_Through_Notification.png")
            assert False
        element.click()
        time.sleep(3)
        deal.clickhere()
        time.sleep(2)
        deal.statusdropdown()
        time.sleep(2)
        deal.clickreject()
        time.sleep(2)
        deal.reasontoreject(self.reason)
        deal.clickonreject()
        time.sleep(3)

    # @pytest.mark.skip(reason="skipping this Test")

    @pytest.mark.regression
    @pytest.mark.run(order=79)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_deal_Create_Verify_Relation_Manager_and_Approve_Check_Relation_manager(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Login Test****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Super admin Login successfully **********")
        deal = dealregistration(driver)
        deal.clickdealtab()
        deal.clickonnewdeal()
        deal.networkcomapnyselect(self.comapanyname)
        deal.companydetails(self.nameofcompany)
        deal.adreesdetails(self.address)
        deal.countrydropdown()
        deal.selectcountry()
        deal.statedropdown()
        deal.selectstate()
        time.sleep(1)
        deal.citydropdown()
        deal.selectcity()
        deal.departmentname(self.departmentfive)
        deal.contactname(self.contname)
        deal.contactemail(self.contemail)
        deal.selectenterprise()
        deal.accountmanager(self.nameofmanager)
        deal.relationmanager()
        deal.enteropportunity(self.oppdetails)
        deal.dealvalue(self.valueofdeal)
        deal.savedeal()
        deal.okaybutton()
        lp.clickLogout()
        lp.setUserName(self.username2)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal.clickdealtab()
        deal.networkdeals()
        deal.pendingdeals()
        deal.dealcompany()
        deal.selectnewdeal()
        lp.clickLogout()
        lp.setUserName(self.username1)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal.clickdealtab()
        deal.networkdeals()
        deal.pendingdeals()
        deal.dealcompany()
        deal.selectnewdeal()
        deal.statusdropdown()
        deal.approved()
        deal.confirmtoapprove()
        time.sleep(3)
        if "Deal Approved Successfully" in driver.page_source:
            self.logger.info("********** Deal approve test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Deal approve test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_dealapprove.png")
            assert False
        time.sleep(3)
        deal.closeapprovedtab()
        lp.clickLogout()
        lp.setUserName(self.username2)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal.clickdealtab()
        deal.networkdeals()
        deal.pendingdeals()
        deal.dealcompany()
        deal.clickonactivedeals()
        deal.selectnewdeal()
        time.sleep(3)
        if "Deal Opportunity Details" in driver.page_source:
            self.logger.info("********** Deal details test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Deal details test is failed **********")
            driver.save_screenshot(
                ".\\Screenshots\\" + "test_deal_Create_Verify_Relation_Manager_and_Approve_Check_Relation_manager.png")
            assert False
        time.sleep(3)

    # @pytest.mark.skip(reason="skipping this Test")

    @pytest.mark.regression
    @pytest.mark.pspk
    @pytest.mark.run(order=80)
    @pytest.mark.flaky(reruns=3, reruns_delay=2)
    def test_deal_Create_Verify_Relation_Manager_and_Approve_Check_Relation_manager_Verify_My_Deals(self, driver):
        driver.maximize_window()
        self.logger.info("****Opening URL****")
        driver.get(self.baseURL)
        self.logger.info("****Started Login Test****")
        lp = LoginPage(driver)
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        self.logger.info("************* Super admin Login successfully **********")
        deal = dealregistration(driver)
        deal.clickdealtab()
        deal.clickonnewdeal()
        deal.networkcomapnyselect(self.comapanyname)
        deal.companydetails(self.nameofcompany)
        deal.adreesdetails(self.address)
        deal.countrydropdown()
        deal.selectcountry()
        deal.statedropdown()
        deal.selectstate()
        time.sleep(1)
        deal.citydropdown()
        deal.selectcity()
        deal.departmentname(self.departmentsix)
        deal.contactname(self.contname)
        deal.contactemail(self.contemail)
        # deal.contactnumber(self.contnumber)
        deal.selectenterprise()
        deal.accountmanager(self.nameofmanager)
        deal.relationmanager()
        deal.enteropportunity(self.oppdetails)
        # deal.currency(self.currencydetails)
        deal.dealvalue(self.valueofdeal)
        deal.savedeal()
        deal.okaybutton()
        lp.clickLogout()
        lp.setUserName(self.username2)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal.clickdealtab()
        deal.networkdeals()
        deal.pendingdeals()
        deal.searchmydealcompany(self.searchtwo)
        time.sleep(1)
        deal.dealcompany()
        deal.selectnewdeal()
        lp.clickLogout()
        lp.setUserName(self.username1)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal.clickdealtab()
        deal.networkdeals()
        deal.pendingdeals()
        deal.searchmydealcompany(self.searchtwo)
        deal.dealcompany()
        deal.selectnewdeal()
        deal.statusdropdown()
        deal.approved()
        deal.confirmtoapprove()
        time.sleep(3)
        if "Deal Approved Successfully" in driver.page_source:
            self.logger.info("********** Deal approve test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Deal approve test is failed **********")
            driver.save_screenshot(".\\Screenshots\\" + "test_dealapprove.png")
            assert False
        time.sleep(3)
        deal.closeapprovedtab()
        lp.clickLogout()
        lp.setUserName(self.username2)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal.clickdealtab()
        deal.networkdeals()
        deal.pendingdeals()
        deal.searchmydealcompany(self.searchtwo)
        time.sleep(1)
        deal.dealcompany()
        deal.clickonactivedeals()
        deal.selectnewdeal()
        lp.clickLogout()
        lp.setUserName(self.username)
        lp.setPassword(self.password)
        lp.clickLogin()
        deal.clickdealtab()
        time.sleep(1)
        deal.searchmydealcompany(self.search)
        time.sleep(1)
        deal.mydealscompany()
        deal.selectnewdealtwo()
        deal.backdealspage()
        deal.mypendingdeals()
        deal.selectnewdealtwo()
        deal.backdealspage()
        deal.myrejectdeals()
        deal.selectnewdealtwo()
        time.sleep(3)
        if "Deal Opportunity Details" in driver.page_source:
            self.logger.info("********** Deal details test is passed *********")
        else:
            # Log and take a screenshot
            self.logger.error("************** Deal details test is failed **********")
            driver.save_screenshot(
                ".\\Screenshots\\" + "test_deal_Create_Verify_Relation_Manager_and_Approve_Check_Relation_manager_Verify_My_Deals.png")
            assert False
        time.sleep(3)
        deal.backdealspage()
        deal.myexpiredeals()
        time.sleep(3)

    if __name__ == '__main__':
        unittest.main(verbosity=2)



